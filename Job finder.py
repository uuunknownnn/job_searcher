from selenium.webdriver.chrome.service import Service
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import os

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

path = r"C:\Users\t90na\Downloads\chromedriver_win32 (2)\chromedriver.exe"
driver = webdriver.Chrome(path)
os.environ['PATH'] += path
driver.get('https://www.jobs.ps/')    # exploring a job website
driver.implicitly_wait(20)
#driver.maximize_window()

search_bar = driver.find_element_by_css_selector('input[class="header--search-input"]')
search_bar.send_keys("Electrical Engineer", Keys.ENTER)  # searching for a specific vacancy

driver.implicitly_wait(20)

#wait till page completely loaded
# get the company names of results
Job_title = driver.find_elements_by_css_selector(
    'a[class="list-3--title list-3--row"]'
)
comp_names = driver.find_elements_by_css_selector(
    'a[class="list-3--title list-3--row"] div[class="list-3--cell-1 list-3--cell-title-2"] div[class="list--cell--company "]'
)
"""
'a[class="list-3--title list-3--row"] div[class="list-3--cell-1"] span[class="tooltip tooltipstered"]'
"""

city = driver.find_elements_by_xpath('//*[@id="main-wrapper"]/div/div/div/div/div/div/a/div/span')

date_posted = driver.find_elements_by_css_selector(
    'a[class="list-3--title list-3--row"] div[class="list-3--cell-1 list-3--cell-4 align-right"]'
)

companies_list = []
for result in comp_names:
    companies_list.append(result.get_attribute(name="innerText"))

job_list = []
for x in Job_title:
    job_list.append(x.get_attribute("title"))

city_list = []
for z in city:
    city_list.append(z.get_attribute(name="innerText"))

post_date = []
for y in date_posted:
    post_date.append(y.get_attribute(name="innerText"))



### createing new workbook

wb = Workbook()
ws = wb.active
ws.title = "Electrical Engineer Jobs"
ws.append(['Company Name','Job Title','Location', 'Date_Published'])


for row in range(2,len(companies_list)+2):

    ws[('A'+str(row))] = companies_list[row-2]

for row in range(2,len(job_list)+2):  
    ws[('B'+str(row))] = job_list[row-2]

for row in range(2,len(city_list)+2):  
    ws[('C'+str(row))] = city_list[row-2]

for row in range(2,len(post_date)+2):  
    ws[('D'+str(row))] = post_date[row-2]

#maximize the width of the column as per the cell contents
dims = {}
for row in ws.rows:
    for cell in row:
        if cell.value:
            dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))   
for col, value in dims.items():
    ws.column_dimensions[col].width = value


wb.save('mynewWorkbook.xlsx')

time.sleep(3)
# quiting the driver
driver.quit()
